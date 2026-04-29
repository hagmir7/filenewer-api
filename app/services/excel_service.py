"""
Excel service functions.
"""

import io
import re
import logging

import pandas as pd

from .csv_service import sanitize_name

logger = logging.getLogger(__name__)


def excel_to_csv(source, sheet_name=None, separator: str = ",") -> dict:
    """
    Convert Excel file (.xlsx / .xls) → CSV string(s).

    Args:
        source     : uploaded file object or file path
        sheet_name : specific sheet name or index (None = all sheets)
        separator  : column separator (default: ',')

    Returns:
        {
            'sheets': {
                'Sheet1': 'csv string...',
                'Sheet2': 'csv string...',
            },
            'total_sheets': 2,
        }
    """
    # ── Read Excel ────────────────────────────────
    if hasattr(source, "read"):
        raw = source.read()
        xls = pd.ExcelFile(io.BytesIO(raw))  # ← BytesIO for binary
    else:
        xls = pd.ExcelFile(source)

    available_sheets = xls.sheet_names

    # ── Decide which sheets to convert ────────────
    if sheet_name is not None:
        # single sheet by name or index
        if isinstance(sheet_name, int):
            if sheet_name >= len(available_sheets):
                raise ValueError(
                    f"Sheet index {sheet_name} out of range. "
                    f"Available: 0-{len(available_sheets)-1}"
                )
            sheet_name = available_sheets[sheet_name]

        if sheet_name not in available_sheets:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. "
                f"Available sheets: {available_sheets}"
            )
        sheets_to_convert = [sheet_name]
    else:
        sheets_to_convert = available_sheets

    # ── Convert each sheet → CSV ──────────────────
    result = {}
    for name in sheets_to_convert:
        df = pd.read_excel(xls, sheet_name=name, engine=None)

        # Clean column names
        df.columns = [sanitize_name(str(c)) for c in df.columns]

        # Replace NaN with empty string for clean CSV
        df = df.fillna("")

        result[name] = df.to_csv(index=False, sep=separator)

    return {
        "sheets": result,
        "total_sheets": len(result),
        "sheet_names": list(result.keys()),
    }


def excel_to_markdown(
    source,
    sheet_name: str = None,
    max_rows: int = 1000,
    include_stats: bool = True,
    include_toc: bool = True,
    column_align: str = "left",
    encoding: str = "utf-8",
) -> dict:
    """
    Convert Excel (.xlsx/.xls) → Markdown text.

    Args:
        source        : uploaded file object OR raw bytes
        sheet_name    : specific sheet name (None = all sheets)
        max_rows      : max rows per sheet             (default: 1000)
        include_stats : include sheet stats header     (default: True)
        include_toc   : include table of contents      (default: True)
        column_align  : left | center | right          (default: left)
        encoding      : output encoding               (default: utf-8)

    Returns:
        {
            'markdown'     : str,
            'sheets'       : list,
            'total_sheets' : int,
            'total_tables' : int,
            'word_count'   : int,
            'char_count'   : int,
        }
    """
    import math

    # ── Read source ───────────────────────────────
    if hasattr(source, "read"):
        raw = source.read()
        filename = getattr(source, "name", "document.xlsx")
    elif isinstance(source, bytes):
        raw = source
        filename = "document.xlsx"
    else:
        raise ValueError("source must be a file object or bytes.")

    if not raw:
        raise ValueError("Empty file.")

    # ── Open Excel ────────────────────────────────
    try:
        xls = pd.ExcelFile(io.BytesIO(raw))
    except Exception as e:
        raise ValueError(f"Cannot open Excel file: {e}")

    available_sheets = xls.sheet_names

    # ── Select sheets ──────────────────────────────
    if sheet_name is not None:
        if sheet_name not in available_sheets:
            raise ValueError(
                f'Sheet "{sheet_name}" not found. ' f"Available: {available_sheets}"
            )
        sheets_to_convert = [sheet_name]
    else:
        sheets_to_convert = available_sheets

    # ── Column alignment ───────────────────────────
    align_map = {
        "left": ":---",
        "center": ":---:",
        "right": "---:",
    }
    align_sep = align_map.get(column_align, ":---")

    # ── Helper: value formatter ────────────────────
    def format_cell(value) -> str:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return ""
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            return str(round(value, 6))
        return str(value).strip()

    def escape_pipe(text: str) -> str:
        return text.replace("|", "\\|").replace("\n", " ").replace("\r", "")

    # ── Process each sheet ─────────────────────────
    md_sections = []
    sheets_info = []
    total_tables = 0

    for sheet in sheets_to_convert:
        try:
            df = pd.read_excel(
                xls,
                sheet_name=sheet,
                nrows=max_rows,
                dtype=str,
                keep_default_na=False,
            )
        except Exception as e:
            md_sections.append(f"## {sheet}\n\n> Error reading sheet: {e}\n")
            continue

        # ── Clean ──────────────────────────────────
        df.columns = [str(c).strip() for c in df.columns]
        df = df.fillna("")

        rows = len(df)
        cols = len(df.columns)
        total_tables += 1

        # ── Sheet stats ────────────────────────────
        sheets_info.append(
            {
                "name": sheet,
                "rows": rows,
                "cols": cols,
            }
        )

        section_lines = []

        # ── Sheet heading ──────────────────────────
        section_lines.append(f"## {sheet}")
        section_lines.append("")

        # ── Stats block ────────────────────────────
        if include_stats:
            section_lines.append(
                f"> **Rows:** {rows} | " f"**Columns:** {cols} | " f"**Sheet:** {sheet}"
            )
            section_lines.append("")

        # ── Empty sheet ────────────────────────────
        if rows == 0:
            section_lines.append("*This sheet is empty.*")
            section_lines.append("")
            md_sections.append("\n".join(section_lines))
            continue

        # ── Build markdown table ───────────────────
        # Headers
        headers = [escape_pipe(format_cell(c)) for c in df.columns]
        header_row = "| " + " | ".join(headers) + " |"
        sep_row = "| " + " | ".join([align_sep] * cols) + " |"

        section_lines.append(header_row)
        section_lines.append(sep_row)

        # Data rows
        for _, row in df.iterrows():
            cells = [escape_pipe(format_cell(v)) for v in row]
            data_row = "| " + " | ".join(cells) + " |"
            section_lines.append(data_row)

        section_lines.append("")

        # ── Column stats ──────────────────────────
        if include_stats:
            section_lines.append("### Column Statistics")
            section_lines.append("")
            section_lines.append("| Column | Type | Unique | Empty |")
            section_lines.append("| :--- | :--- | ---: | ---: |")

            for col in df.columns:
                series = df[col]
                numeric = pd.to_numeric(series, errors="coerce")
                is_numeric = numeric.notna().sum() > len(series) * 0.5
                col_type = "Numeric" if is_numeric else "Text"
                unique = series.nunique()
                empty = (series.str.strip() == "").sum()

                stat_row = (
                    f"| {escape_pipe(col)} "
                    f"| {col_type} "
                    f"| {unique} "
                    f"| {empty} |"
                )
                section_lines.append(stat_row)

                if is_numeric and not numeric.isna().all():
                    section_lines.append(
                        f"| ↳ *min: {round(float(numeric.min()), 2)}, "
                        f"max: {round(float(numeric.max()), 2)}, "
                        f"mean: {round(float(numeric.mean()), 2)}, "
                        f"sum: {round(float(numeric.sum()), 2)}* "
                        f"| | | |"
                    )

            section_lines.append("")

        md_sections.append("\n".join(section_lines))

    # ── Build full markdown ────────────────────────
    parts = []

    # ── Document header ────────────────────────────
    parts.append(f'# {filename.replace(".xlsx","").replace(".xls","")}')
    parts.append("")
    parts.append(
        f"> **File:** {filename} | "
        f"**Sheets:** {len(sheets_to_convert)} | "
        f'**Total rows:** {sum(s["rows"] for s in sheets_info)}'
    )
    parts.append("")

    # ── Table of contents ─────────────────────────
    if include_toc and len(sheets_to_convert) > 1:
        parts.append("## Contents")
        parts.append("")
        for s in sheets_info:
            slug = s["name"].lower().replace(" ", "-")
            slug = "".join(c for c in slug if c.isalnum() or c == "-")
            parts.append(
                f'- [{s["name"]}](#{slug}) ' f'— {s["rows"]} rows × {s["cols"]} cols'
            )
        parts.append("")
        parts.append("---")
        parts.append("")

    # ── Sheet sections ─────────────────────────────
    parts.extend(md_sections)

    # ── Join ──────────────────────────────────────
    import re

    markdown = "\n".join(parts)
    markdown = re.sub(r"\n{3,}", "\n\n", markdown).strip()

    word_count = len(markdown.split())
    char_count = len(markdown)

    return {
        "markdown": markdown,
        "sheets": sheets_info,
        "total_sheets": len(sheets_info),
        "total_tables": total_tables,
        "word_count": word_count,
        "char_count": char_count,
        "encoding": encoding,
    }


def _safe_auto_fit(ws, min_width: int = 8, max_width: int = 60):
    """
    Auto-fit column widths.
    Safely skips MergedCell objects which have no column_letter attribute.
    Call this instead of any loop using col_cells[0].column_letter.
    """
    from openpyxl.cell.cell import MergedCell

    col_widths = {}

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            try:
                col_letter = cell.column_letter
            except AttributeError:
                continue
            if not col_letter:
                continue

            length = len(str(cell.value)) if cell.value is not None else 0
            col_widths[col_letter] = max(
                col_widths.get(col_letter, min_width),
                min(length + 4, max_width),
            )

    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width


def markdown_to_excel(
    source,
    filename      : str  = 'document.md',
    sheet_name    : str  = None,
    encoding      : str  = 'utf-8',
    include_stats : bool = True,
) -> bytes:
    """
    Convert Markdown (.md) → Excel (.xlsx).

    Extracts:
        - Tables      → dedicated sheet per table
        - Code blocks → Code sheet
        - Headings    → Structure sheet
        - Lists       → Lists sheet
        - Stats       → Summary sheet

    Args:
        source        : file object | raw markdown string | bytes
        filename      : original filename
        sheet_name    : default sheet name            (default: Sheet1)
        encoding      : input encoding               (default: utf-8)
        include_stats : include summary sheet         (default: True)

    Returns:
        Raw bytes of the generated .xlsx file
    """
    from openpyxl        import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import re

    # ── Read source ───────────────────────────────
    if hasattr(source, 'read'):
        raw = source.read()
        if isinstance(raw, bytes):
            raw = raw.decode(encoding, errors='replace')
    elif isinstance(source, bytes):
        raw = source.decode(encoding, errors='replace')
    elif isinstance(source, str):
        raw = source
    else:
        raise ValueError('source must be a string, bytes, or file object.')

    if not raw.strip():
        raise ValueError('Empty input.')

    # ── Style constants ────────────────────────────
    BLUE       = '2E75B6'
    LIGHT_BLUE = 'DCE6F1'
    DARK_BLUE  = '1F4E79'
    GREEN      = '70AD47'
    ORANGE     = 'ED7D31'
    GREY       = 'F2F2F2'
    WHITE      = 'FFFFFF'

    thin_border = Border(
        left  =Side(style='thin'),
        right =Side(style='thin'),
        top   =Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # ── Style helpers ──────────────────────────────
    def header_style(cell, color=BLUE):
        cell.font      = Font(bold=True, color=WHITE, size=11)
        cell.fill      = PatternFill(fill_type='solid', fgColor=color)
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
        cell.border = thin_border

    def data_style(cell, even_row=False, bold=False):
        cell.fill      = PatternFill(
            fill_type='solid',
            fgColor=LIGHT_BLUE if even_row else WHITE,
        )
        cell.font      = Font(bold=bold, size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border    = thin_border

    # ── Inline markdown stripper ───────────────────
    def strip_md_inline(text: str) -> str:
        text = re.sub(r'\*\*\*(.+?)\*\*\*', r'\1', text)
        text = re.sub(r'\*\*(.+?)\*\*',     r'\1', text)
        text = re.sub(r'__(.+?)__',          r'\1', text)
        text = re.sub(r'\*(.+?)\*',          r'\1', text)
        text = re.sub(r'_(.+?)_',            r'\1', text)
        text = re.sub(r'~~(.+?)~~',          r'\1', text)
        text = re.sub(r'`(.+?)`',            r'\1', text)
        text = re.sub(r'\[(.+?)\]\(.+?\)',   r'\1', text)
        text = re.sub(r'<[^>]+>',            '',    text)
        return text.strip()

    # ─────────────────────────────────────────────
    # Parse Markdown
    # ─────────────────────────────────────────────
    lines               = raw.splitlines()
    total               = len(lines)
    i                   = 0
    tables              = []
    code_blocks         = []
    headings            = []
    list_items          = []
    paragraphs          = []
    current_table_title = ''

    while i < total:
        line     = lines[i]
        stripped = line.strip()

        # ── Empty line ─────────────────────────────
        if not stripped:
            current_table_title = ''
            i += 1
            continue

        # ── Fenced code block ──────────────────────
        if stripped.startswith('```') or stripped.startswith('~~~'):
            fence      = stripped[:3]
            language   = stripped[3:].strip()
            i         += 1
            code_lines = []
            while i < total and not lines[i].strip().startswith(fence):
                code_lines.append(lines[i])
                i += 1
            code_blocks.append({
                'language': language or 'plain',
                'code'    : '\n'.join(code_lines),
            })
            i += 1
            continue

        # ── ATX Heading ────────────────────────────
        m = re.match(r'^(#{1,6})\s+(.+)$', stripped)
        if m:
            level = len(m.group(1))
            text  = strip_md_inline(m.group(2).rstrip('#').strip())
            headings.append({'level': level, 'text': text})
            current_table_title = text
            i += 1
            continue

        # ── Setext Heading ─────────────────────────
        if i + 1 < total:
            next_line = lines[i + 1].strip()
            if next_line and set(next_line) == {'='}:
                text = strip_md_inline(stripped)
                headings.append({'level': 1, 'text': text})
                current_table_title = text
                i += 2
                continue
            if next_line and set(next_line) == {'-'} and len(next_line) >= 2:
                text = strip_md_inline(stripped)
                headings.append({'level': 2, 'text': text})
                current_table_title = text
                i += 2
                continue

        # ── Table ──────────────────────────────────
        if stripped.startswith('|') and '|' in stripped[1:]:
            table_lines = []
            while i < total and lines[i].strip().startswith('|'):
                table_lines.append(lines[i].strip())
                i += 1

            headers  = []
            rows     = []
            is_first = True

            for tl in table_lines:
                if re.match(
                    r'^\|?[\s:]*-+[\s:]*(\|[\s:]*-+[\s:]*)*\|?$', tl
                ):
                    continue
                cells = [
                    strip_md_inline(c.strip())
                    for c in tl.strip('|').split('|')
                ]
                if is_first:
                    headers  = cells
                    is_first = False
                else:
                    rows.append(cells)

            if headers:
                tables.append({
                    'title'  : current_table_title or f'Table {len(tables)+1}',
                    'headers': headers,
                    'rows'   : rows,
                })
            continue

        # ── Bullet list ────────────────────────────
        m = re.match(r'^(\s*)[-*+]\s+(\[[ xX]\]\s+)?(.+)$', line)
        if m:
            level   = len(m.group(1)) // 2
            checked = None
            text    = strip_md_inline(m.group(3))
            if m.group(2):
                checked = m.group(2).strip()[1].lower() == 'x'
            list_items.append({
                'type'   : 'bullet',
                'level'  : level,
                'text'   : text,
                'checked': checked,
            })
            i += 1
            continue

        # ── Numbered list ──────────────────────────
        m = re.match(r'^(\s*)\d+[.)]\s+(.+)$', line)
        if m:
            level = len(m.group(1)) // 3
            text  = strip_md_inline(m.group(2))
            list_items.append({
                'type' : 'number',
                'level': level,
                'text' : text,
            })
            i += 1
            continue

        # ── Horizontal rule ────────────────────────
        if re.match(r'^[-*_]{3,}$', stripped.replace(' ', '')):
            i += 1
            continue

        # ── Regular paragraph ──────────────────────
        if stripped and not stripped.startswith('>'):
            paragraphs.append(strip_md_inline(stripped))

        i += 1

    # ─────────────────────────────────────────────
    # Build Excel Workbook
    # ─────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ────────────────────────────────────────────────
    # Sheet: Summary
    # ────────────────────────────────────────────────
    if include_stats:
        ws_summary = wb.create_sheet('Summary')
        ws_summary.sheet_properties.tabColor = BLUE

        # Title (merged A1:C1)
        ws_summary['A1'] = f'Markdown Document: {filename}'
        ws_summary['A1'].font = Font(bold=True, size=14, color=DARK_BLUE)
        ws_summary.merge_cells('A1:C1')
        ws_summary.row_dimensions[1].height = 30

        stats = [
            ('Property',       'Value',         'Details'),
            ('Filename',        filename,         ''),
            ('Total Headings',  len(headings),    'H1–H6 headings found'),
            ('Tables Found',    len(tables),
             f'{sum(len(t["rows"]) for t in tables)} total rows'),
            ('Code Blocks',     len(code_blocks),
             ', '.join(set(c['language'] for c in code_blocks)) or 'none'),
            ('List Items',      len(list_items),
             f'{sum(1 for l in list_items if l["type"]=="bullet")} bullets, '
             f'{sum(1 for l in list_items if l["type"]=="number")} numbered'),
            ('Paragraphs',      len(paragraphs),  ''),
            ('Total Words',     len(raw.split()), ''),
            ('Total Chars',     len(raw),          ''),
        ]

        for row_idx, row_data in enumerate(stats, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 2:
                    header_style(cell)
                else:
                    data_style(cell, even_row=(row_idx % 2 == 0))

        # Safe auto-fit — skips merged B1/C1
        _safe_auto_fit(ws_summary)
        ws_summary.freeze_panes = 'A3'

    # ────────────────────────────────────────────────
    # Sheet: Structure (Headings)
    # ────────────────────────────────────────────────
    if headings:
        ws_struct = wb.create_sheet('Structure')
        ws_struct.sheet_properties.tabColor = DARK_BLUE

        for col_idx, h in enumerate(
            ['Level', 'Heading', 'Indent'], start=1
        ):
            header_style(ws_struct.cell(row=1, column=col_idx, value=h))

        for row_idx, h in enumerate(headings, start=2):
            indent = '  ' * (h['level'] - 1) + h['text']
            even   = row_idx % 2 == 0

            lv = ws_struct.cell(row=row_idx, column=1, value=f'H{h["level"]}')
            lv.font      = Font(bold=True, color=WHITE, size=10)
            lv.fill      = PatternFill(
                fill_type='solid',
                fgColor=BLUE if h['level'] <= 2 else DARK_BLUE,
            )
            lv.alignment = Alignment(horizontal='center', vertical='center')
            lv.border    = thin_border

            h2 = ws_struct.cell(row=row_idx, column=2, value=h['text'])
            h2.font = Font(bold=h['level'] <= 2, size=10)
            data_style(h2, even_row=even)

            data_style(
                ws_struct.cell(row=row_idx, column=3, value=indent),
                even_row=even,
            )

        # Safe auto-fit
        _safe_auto_fit(ws_struct)
        ws_struct.freeze_panes = 'A2'

    # ────────────────────────────────────────────────
    # Sheets: One per table
    # ────────────────────────────────────────────────
    for t_idx, table in enumerate(tables, start=1):
        safe_name   = re.sub(r'[^\w\s]', '', table['title'])[:28].strip()
        sheet_label = (safe_name or f'Table_{t_idx}')[:31]

        existing = [ws.title for ws in wb.worksheets]
        if sheet_label in existing:
            sheet_label = f'{sheet_label[:27]}_{t_idx}'

        ws_tbl = wb.create_sheet(sheet_label)
        ws_tbl.sheet_properties.tabColor = GREEN

        # Title row (merged)
        title_cell      = ws_tbl.cell(row=1, column=1, value=table['title'])
        title_cell.font = Font(bold=True, size=13, color=DARK_BLUE)
        if table['headers']:
            ws_tbl.merge_cells(
                start_row   =1, end_row=1,
                start_column=1,
                end_column  =len(table['headers']),
            )
        ws_tbl.row_dimensions[1].height = 25

        # Headers
        for col_idx, header in enumerate(table['headers'], start=1):
            header_style(ws_tbl.cell(row=2, column=col_idx, value=header))
        ws_tbl.row_dimensions[2].height = 22
        ws_tbl.freeze_panes = 'A3'

        # Data rows
        for row_idx, row_data in enumerate(table['rows'], start=3):
            num_cols = len(table['headers'])
            for col_idx in range(num_cols):
                raw_val = row_data[col_idx] if col_idx < len(row_data) else ''
                value   = raw_val
                try:
                    value = int(raw_val) if '.' not in str(raw_val) \
                        else float(raw_val)
                except (ValueError, TypeError):
                    pass

                cell = ws_tbl.cell(
                    row=row_idx, column=col_idx + 1, value=value
                )
                data_style(cell, even_row=(row_idx % 2 == 0))
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(
                        horizontal='right', vertical='center'
                    )

        # Totals row
        if table['rows']:
            total_row = len(table['rows']) + 3

            tc0 = ws_tbl.cell(row=total_row, column=1, value='TOTAL')
            tc0.font      = Font(bold=True, color=WHITE)
            tc0.fill      = PatternFill(fill_type='solid', fgColor=BLUE)
            tc0.alignment = Alignment(horizontal='center')
            tc0.border    = thin_border

            for col_idx in range(1, len(table['headers'])):
                numeric_vals = []
                for rd in table['rows']:
                    if col_idx < len(rd):
                        try:
                            numeric_vals.append(float(rd[col_idx]))
                        except (ValueError, TypeError):
                            pass

                tc = ws_tbl.cell(row=total_row, column=col_idx + 1)
                if numeric_vals:
                    tc.value     = sum(numeric_vals)
                    tc.font      = Font(bold=True, color=WHITE)
                    tc.alignment = Alignment(
                        horizontal='right', vertical='center'
                    )
                tc.fill   = PatternFill(fill_type='solid', fgColor=BLUE)
                tc.border = thin_border

        # Safe auto-fit — skips merged title row cells
        _safe_auto_fit(ws_tbl)

    # ────────────────────────────────────────────────
    # Sheet: Lists
    # ────────────────────────────────────────────────
    if list_items:
        ws_list = wb.create_sheet('Lists')
        ws_list.sheet_properties.tabColor = ORANGE

        for col_idx, h in enumerate(
            ['#', 'Type', 'Level', 'Item', 'Status'], start=1
        ):
            header_style(
                ws_list.cell(row=1, column=col_idx, value=h),
                color=ORANGE,
            )
        ws_list.freeze_panes = 'A2'

        for row_idx, item in enumerate(list_items, start=2):
            even = row_idx % 2 == 0

            nc = ws_list.cell(row=row_idx, column=1, value=row_idx - 1)
            data_style(nc, even_row=even)
            nc.alignment = Alignment(horizontal='center', vertical='center')

            type_label = '• Bullet' if item['type'] == 'bullet' else '1. Number'
            data_style(
                ws_list.cell(row=row_idx, column=2, value=type_label),
                even_row=even,
            )

            lvc = ws_list.cell(row=row_idx, column=3, value=item['level'])
            data_style(lvc, even_row=even)
            lvc.alignment = Alignment(horizontal='center', vertical='center')

            indent_text = '  ' * item['level'] + item['text']
            data_style(
                ws_list.cell(row=row_idx, column=4, value=indent_text),
                even_row=even,
                bold=(item['level'] == 0),
            )

            status  = ''
            checked = item.get('checked')
            if checked is not None:
                status = '☑ Done' if checked else '☐ Todo'
            st = ws_list.cell(row=row_idx, column=5, value=status)
            data_style(st, even_row=even)
            if status == '☑ Done':
                st.font = Font(color='70AD47', bold=True, size=10)
            elif status == '☐ Todo':
                st.font = Font(color='ED7D31', bold=True, size=10)

        # Safe auto-fit
        _safe_auto_fit(ws_list)

    # ────────────────────────────────────────────────
    # Sheet: Code
    # ────────────────────────────────────────────────
    if code_blocks:
        ws_code = wb.create_sheet('Code')
        ws_code.sheet_properties.tabColor = '404040'

        for col_idx, h in enumerate(
            ['#', 'Language', 'Code', 'Lines'], start=1
        ):
            header_style(
                ws_code.cell(row=1, column=col_idx, value=h),
                color='404040',
            )
        ws_code.freeze_panes = 'A2'

        for row_idx, block in enumerate(code_blocks, start=2):
            even = row_idx % 2 == 0

            nc = ws_code.cell(row=row_idx, column=1, value=row_idx - 1)
            data_style(nc, even_row=even)
            nc.alignment = Alignment(horizontal='center', vertical='center')

            data_style(
                ws_code.cell(row=row_idx, column=2, value=block['language']),
                even_row=even,
                bold=True,
            )

            cc           = ws_code.cell(
                row=row_idx, column=3, value=block['code']
            )
            cc.font      = Font(name='Courier New', size=9)
            cc.fill      = PatternFill(fill_type='solid', fgColor=GREY)
            cc.alignment = Alignment(vertical='top', wrap_text=True)
            cc.border    = thin_border
            ws_code.row_dimensions[row_idx].height = min(
                15 * block['code'].count('\n') + 15, 200
            )

            lc = ws_code.cell(
                row=row_idx, column=4,
                value=block['code'].count('\n') + 1,
            )
            data_style(lc, even_row=even)
            lc.alignment = Alignment(horizontal='center', vertical='center')

        # Fixed column widths (no auto-fit — code column is wide)
        ws_code.column_dimensions['A'].width = 5
        ws_code.column_dimensions['B'].width = 15
        ws_code.column_dimensions['C'].width = 80
        ws_code.column_dimensions['D'].width = 8

    # ────────────────────────────────────────────────
    # Sheet: Content (paragraphs)
    # ────────────────────────────────────────────────
    if paragraphs:
        ws_content = wb.create_sheet('Content')
        ws_content.sheet_properties.tabColor = GREEN

        for col_idx, h in enumerate(['#', 'Paragraph'], start=1):
            header_style(
                ws_content.cell(row=1, column=col_idx, value=h),
                color=GREEN,
            )
        ws_content.freeze_panes = 'A2'

        for row_idx, para in enumerate(paragraphs, start=2):
            even = row_idx % 2 == 0

            nc = ws_content.cell(row=row_idx, column=1, value=row_idx - 1)
            data_style(nc, even_row=even)
            nc.alignment = Alignment(horizontal='center', vertical='center')

            pc           = ws_content.cell(row=row_idx, column=2, value=para)
            pc.font      = Font(size=10)
            pc.fill      = PatternFill(
                fill_type='solid',
                fgColor=LIGHT_BLUE if even else WHITE,
            )
            pc.alignment = Alignment(vertical='top', wrap_text=True)
            pc.border    = thin_border

        # Fixed widths for content sheet
        ws_content.column_dimensions['A'].width = 5
        ws_content.column_dimensions['B'].width = 100

    # ── Fallback ───────────────────────────────────
    if not wb.worksheets:
        ws = wb.create_sheet('Content')
        ws['A1'] = 'No structured content found in Markdown.'
        ws['A1'].font = Font(size=11, color='FF0000')

    # ── Serialize ─────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def sql_to_excel(
    source,
    filename: str = "query.sql",
    encoding: str = "utf-8",
    include_stats: bool = True,
    include_schema: bool = True,
) -> bytes:
    """
    Parse SQL (DDL + INSERT statements) → Excel (.xlsx).

    Handles:
        - CREATE TABLE  → schema sheet
        - INSERT INTO   → data sheet per table
        - Multiple tables → one sheet per table
        - Column types   → type detection
        - Statistics     → summary sheet

    Args:
        source         : file object | raw SQL string | bytes
        filename       : original filename
        encoding       : input encoding               (default: utf-8)
        include_stats  : include summary sheet        (default: True)
        include_schema : include schema sheet         (default: True)

    Returns:
        Raw bytes of the generated .xlsx file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import re

    # ── Read source ───────────────────────────────
    if hasattr(source, "read"):
        raw = source.read()
        if isinstance(raw, bytes):
            raw = raw.decode(encoding, errors="replace")
    elif isinstance(source, bytes):
        raw = source.decode(encoding, errors="replace")
    elif isinstance(source, str):
        raw = source
    else:
        raise ValueError("source must be a string, bytes, or file object.")

    if not raw.strip():
        raise ValueError("Empty input.")

    # ── Style constants ────────────────────────────
    BLUE = "2E75B6"
    LIGHT_BLUE = "DCE6F1"
    DARK_BLUE = "1F4E79"
    GREEN = "70AD47"
    ORANGE = "ED7D31"
    PURPLE = "7030A0"
    WHITE = "FFFFFF"
    GREY = "F2F2F2"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def header_style(cell, color=BLUE):
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor=color)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = thin_border

    def data_style(cell, even_row=False, bold=False, align="left"):
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=LIGHT_BLUE if even_row else WHITE,
        )
        cell.font = Font(bold=bold, size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = thin_border

    # ── Safe auto-fit ──────────────────────────────
    def safe_auto_fit(ws, min_w=8, max_w=60):
        from openpyxl.cell.cell import MergedCell

        col_widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                try:
                    col_letter = cell.column_letter
                except AttributeError:
                    continue
                if not col_letter:
                    continue
                length = len(str(cell.value)) if cell.value is not None else 0
                col_widths[col_letter] = max(
                    col_widths.get(col_letter, min_w),
                    min(length + 4, max_w),
                )
        for letter, width in col_widths.items():
            ws.column_dimensions[letter].width = width

    # ─────────────────────────────────────────────
    # Parse SQL
    # ─────────────────────────────────────────────

    # ── Remove comments ────────────────────────────
    sql = re.sub(r"--[^\n]*", "", raw)  # single line
    sql = re.sub(r"/\*.*?\*/", "", sql, flags=re.DOTALL)  # block
    sql = re.sub(r"\s+", " ", sql).strip()

    # ── Parse CREATE TABLE ────────────────────────
    schemas = {}  # { table_name: [ {col, type, constraints} ] }

    create_pattern = re.compile(
        r"CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?"
        r'[`"\[]?(\w+)[`"\]]?\s*\((.+?)\)\s*;',
        re.IGNORECASE | re.DOTALL,
    )

    for match in create_pattern.finditer(sql):
        table_name = match.group(1).strip()
        cols_raw = match.group(2).strip()

        columns = []
        for col_def in _split_columns(cols_raw):
            col_def = col_def.strip()
            if not col_def:
                continue

            # Skip constraints
            upper = col_def.upper()
            if any(
                upper.startswith(k)
                for k in (
                    "PRIMARY",
                    "FOREIGN",
                    "UNIQUE",
                    "CHECK",
                    "INDEX",
                    "KEY",
                    "CONSTRAINT",
                )
            ):
                continue

            # Parse column name and type
            parts = col_def.split()
            if len(parts) >= 2:
                col_name = parts[0].strip('`"\[]')
                col_type = parts[1].strip('`"\[],()')
                constraints = " ".join(parts[2:]).upper()
                columns.append(
                    {
                        "name": col_name,
                        "type": col_type,
                        "nullable": "NOT NULL" not in constraints,
                        "primary_key": "PRIMARY" in constraints,
                        "default": _extract_default(col_def),
                        "constraints": constraints,
                    }
                )

        if columns:
            schemas[table_name.upper()] = columns

    # ── Parse INSERT INTO ──────────────────────────
    tables = {}  # { table_name: { 'columns': [], 'rows': [] } }

    insert_pattern = re.compile(
        r'INSERT\s+INTO\s+[`"\[]?(\w+)[`"\]]?\s*'
        r"(?:\(([^)]+)\))?\s*VALUES\s*(.+?)(?=INSERT\s+INTO|$)",
        re.IGNORECASE | re.DOTALL,
    )

    for match in insert_pattern.finditer(sql):
        table_name = match.group(1).strip().upper()
        cols_str = match.group(2)
        values_str = match.group(3).strip()

        # Parse column names
        if cols_str:
            columns = [c.strip().strip('`"\[]') for c in cols_str.split(",")]
        else:
            # Use schema columns if available
            columns = [c["name"] for c in schemas.get(table_name, [])]

        if table_name not in tables:
            tables[table_name] = {"columns": columns, "rows": []}
        elif not tables[table_name]["columns"] and columns:
            tables[table_name]["columns"] = columns

        # Parse VALUES
        rows = _parse_values(values_str)
        tables[table_name]["rows"].extend(rows)

    # ─────────────────────────────────────────────
    # Build Excel Workbook
    # ─────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # ────────────────────────────────────────────────
    # Sheet: Summary
    # ────────────────────────────────────────────────
    if include_stats:
        ws_sum = wb.create_sheet("Summary")
        ws_sum.sheet_properties.tabColor = BLUE

        ws_sum["A1"] = f"SQL File: {filename}"
        ws_sum["A1"].font = Font(bold=True, size=14, color=DARK_BLUE)
        ws_sum.merge_cells("A1:D1")
        ws_sum.row_dimensions[1].height = 30

        stats = [
            ("Property", "Value", "Details", ""),
            ("Filename", filename, "", ""),
            ("Tables Found", len(tables), f'{", ".join(tables.keys()) or "none"}', ""),
            (
                "Total Rows",
                sum(len(t["rows"]) for t in tables.values()),
                "across all tables",
                "",
            ),
            (
                "Schemas Parsed",
                len(schemas),
                f'{", ".join(schemas.keys()) or "none"}',
                "",
            ),
            ("SQL Size", f"{round(len(raw)/1024, 2)} KB", "", ""),
        ]

        for r_idx, row_data in enumerate(stats, start=2):
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws_sum.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 2:
                    header_style(cell)
                else:
                    data_style(cell, even_row=(r_idx % 2 == 0))

        safe_auto_fit(ws_sum)
        ws_sum.freeze_panes = "A3"

    # ────────────────────────────────────────────────
    # Sheet: Schema
    # ────────────────────────────────────────────────
    if include_schema and schemas:
        ws_schema = wb.create_sheet("Schema")
        ws_schema.sheet_properties.tabColor = PURPLE

        headers = ["Table", "Column", "Type", "Nullable", "Primary Key", "Default"]
        for c_idx, h in enumerate(headers, start=1):
            header_style(
                ws_schema.cell(row=1, column=c_idx, value=h),
                color=PURPLE,
            )
        ws_schema.freeze_panes = "A2"

        r_idx = 2
        for table_name, columns in schemas.items():
            for col in columns:
                even = r_idx % 2 == 0

                # Table name
                tc = ws_schema.cell(row=r_idx, column=1, value=table_name)
                tc.font = Font(bold=True, size=10, color=DARK_BLUE)
                tc.fill = PatternFill(
                    fill_type="solid",
                    fgColor=LIGHT_BLUE if even else WHITE,
                )
                tc.alignment = Alignment(vertical="center")
                tc.border = thin_border

                data_style(
                    ws_schema.cell(row=r_idx, column=2, value=col["name"]),
                    even_row=even,
                    bold=col["primary_key"],
                )

                type_cell = ws_schema.cell(row=r_idx, column=3, value=col["type"])
                type_cell.font = Font(
                    name="Courier New",
                    size=9,
                    color="C0392B",
                )
                type_cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=LIGHT_BLUE if even else WHITE,
                )
                type_cell.alignment = Alignment(vertical="center")
                type_cell.border = thin_border

                # Nullable
                null_val = "✓ Yes" if col["nullable"] else "✗ No"
                null_cell = ws_schema.cell(row=r_idx, column=4, value=null_val)
                null_cell.font = Font(
                    bold=True,
                    size=10,
                    color="70AD47" if col["nullable"] else "C0392B",
                )
                null_cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=LIGHT_BLUE if even else WHITE,
                )
                null_cell.alignment = Alignment(horizontal="center", vertical="center")
                null_cell.border = thin_border

                # Primary Key
                pk_val = "🔑 Yes" if col["primary_key"] else ""
                pk_cell = ws_schema.cell(row=r_idx, column=5, value=pk_val)
                data_style(pk_cell, even_row=even)
                pk_cell.alignment = Alignment(horizontal="center", vertical="center")

                data_style(
                    ws_schema.cell(
                        row=r_idx,
                        column=6,
                        value=col["default"] or "",
                    ),
                    even_row=even,
                )

                r_idx += 1

        safe_auto_fit(ws_schema)

    # ────────────────────────────────────────────────
    # Sheets: One per table (INSERT data)
    # ────────────────────────────────────────────────
    for table_name, table_data in tables.items():
        cols = table_data["columns"]
        rows = table_data["rows"]

        safe_label = table_name[:31]
        existing = [ws.title for ws in wb.worksheets]
        if safe_label in existing:
            safe_label = safe_label[:28] + "_2"

        ws_tbl = wb.create_sheet(safe_label)
        ws_tbl.sheet_properties.tabColor = GREEN

        # ── Title row ──────────────────────────────
        title_cell = ws_tbl.cell(
            row=1,
            column=1,
            value=f"{table_name} ({len(rows)} rows)",
        )
        title_cell.font = Font(bold=True, size=13, color=DARK_BLUE)
        if cols:
            ws_tbl.merge_cells(
                start_row=1,
                end_row=1,
                start_column=1,
                end_column=max(len(cols), 1),
            )
        ws_tbl.row_dimensions[1].height = 25

        # ── Column headers ─────────────────────────
        if not cols and rows:
            cols = [f"col_{i+1}" for i in range(len(rows[0]))]

        # Add schema type hint to header
        schema_cols = {
            c["name"].upper(): c["type"] for c in schemas.get(table_name, [])
        }

        for c_idx, col in enumerate(cols, start=1):
            col_type = schema_cols.get(col.upper(), "")
            header_label = f"{col}\n({col_type})" if col_type else col
            cell = ws_tbl.cell(row=2, column=c_idx, value=header_label)
            header_style(cell)

        ws_tbl.row_dimensions[2].height = 30
        ws_tbl.freeze_panes = "A3"

        # ── Data rows ──────────────────────────────
        for r_idx, row in enumerate(rows, start=3):
            even = r_idx % 2 == 0
            for c_idx, val in enumerate(row, start=1):
                if c_idx > len(cols):
                    break

                # Type coercion
                typed_val = _coerce_sql_value(val)

                cell = ws_tbl.cell(row=r_idx, column=c_idx, value=typed_val)

                align = "right" if isinstance(typed_val, (int, float)) else "left"
                data_style(cell, even_row=even, align=align)

        # ── Totals row for numeric columns ─────────
        if rows:
            total_row = len(rows) + 3

            tot0 = ws_tbl.cell(row=total_row, column=1, value="TOTAL")
            tot0.font = Font(bold=True, color=WHITE)
            tot0.fill = PatternFill(fill_type="solid", fgColor=BLUE)
            tot0.alignment = Alignment(horizontal="center")
            tot0.border = thin_border

            for c_idx in range(1, len(cols)):
                numeric_vals = []
                for row in rows:
                    if c_idx < len(row):
                        v = _coerce_sql_value(row[c_idx])
                        if isinstance(v, (int, float)):
                            numeric_vals.append(v)

                tc = ws_tbl.cell(row=total_row, column=c_idx + 1)
                if numeric_vals:
                    tc.value = sum(numeric_vals)
                    tc.font = Font(bold=True, color=WHITE)
                    tc.alignment = Alignment(horizontal="right", vertical="center")
                tc.fill = PatternFill(fill_type="solid", fgColor=BLUE)
                tc.border = thin_border

        safe_auto_fit(ws_tbl)

    # ── Fallback empty ─────────────────────────────
    if not wb.worksheets:
        ws = wb.create_sheet("Result")
        ws["A1"] = "No CREATE TABLE or INSERT INTO statements found in SQL."
        ws["A1"].font = Font(size=11, color="FF0000")

    # ── Serialize ─────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _split_columns(cols_raw: str) -> list:
    """
    Split CREATE TABLE column definitions respecting parentheses.
    e.g. col1 INT, col2 VARCHAR(255), CHECK(col1 > 0)
    """
    parts = []
    depth = 0
    current = ""

    for ch in cols_raw:
        if ch == "(":
            depth += 1
            current += ch
        elif ch == ")":
            depth -= 1
            current += ch
        elif ch == "," and depth == 0:
            parts.append(current.strip())
            current = ""
        else:
            current += ch

    if current.strip():
        parts.append(current.strip())

    return parts


def _extract_default(col_def: str) -> str:
    """Extract DEFAULT value from column definition."""
    import re

    m = re.search(r"DEFAULT\s+([^\s,]+)", col_def, re.IGNORECASE)
    return m.group(1).strip("'\"") if m else ""


def _parse_values(values_str: str) -> list:
    """
    Parse VALUES clause into list of rows.
    Handles: (1,'text',NULL), (2,'it''s',3.14)
    """
    import re

    rows = []
    # Find all value groups (...), (...), ...
    groups = re.findall(r"\(([^)]*(?:\([^)]*\)[^)]*)*)\)", values_str)

    for group in groups:
        values = _split_values(group)
        rows.append(values)

    return rows


def _split_values(values_str: str) -> list:
    """Split a VALUES group respecting quoted strings."""
    parts = []
    current = ""
    in_str = False
    quote = None

    i = 0
    while i < len(values_str):
        ch = values_str[i]

        if not in_str and ch in ("'", '"'):
            in_str = True
            quote = ch
            current += ch
        elif in_str and ch == quote:
            # Check for escaped quote ''
            if i + 1 < len(values_str) and values_str[i + 1] == quote:
                current += ch + quote
                i += 2
                continue
            else:
                in_str = False
                current += ch
        elif not in_str and ch == ",":
            parts.append(current.strip())
            current = ""
        else:
            current += ch

        i += 1

    if current.strip():
        parts.append(current.strip())

    return parts


def _coerce_sql_value(val: str):
    """Convert SQL string value to Python type."""
    if val is None:
        return None

    val = str(val).strip()

    # NULL
    if val.upper() == "NULL":
        return None

    # Boolean
    if val.upper() in ("TRUE", "1"):
        return True
    if val.upper() in ("FALSE", "0") and val.upper() in ("TRUE", "FALSE"):
        return False

    # Strip quotes
    if (val.startswith("'") and val.endswith("'")) or (
        val.startswith('"') and val.endswith('"')
    ):
        return val[1:-1].replace("''", "'")

    # Integer
    try:
        return int(val)
    except ValueError:
        pass

    # Float
    try:
        return float(val)
    except ValueError:
        pass

    return val

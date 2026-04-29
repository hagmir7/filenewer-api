"""
JSON service functions.
"""

import io
import json
import logging

import pandas as pd

from .csv_service import sanitize_name

logger = logging.getLogger(__name__)


def json_to_csv(source, separator: str = ",") -> str:
    """
    Convert JSON (file, raw text, or dict/list) → CSV string.

    Args:
        source    : uploaded file object | raw JSON string | list | dict
        separator : column separator (default: ',')

    Returns:
        CSV string
    """
    # ── Parse input ───────────────────────────────
    if isinstance(source, (list, dict)):
        data = source  # already parsed
    elif isinstance(source, str):
        data = json.loads(source)  # raw JSON string
    else:
        raw = source.read().decode("utf-8")  # file object
        data = json.loads(raw)

    # ── Handle nested/wrapped JSON ─────────────────
    # e.g. { "data": [ {...}, {...} ] }
    if isinstance(data, dict):
        # find the first key whose value is a list
        for key, val in data.items():
            if isinstance(val, list):
                data = val
                break
        else:
            data = [data]  # single object → wrap in list

    if not isinstance(data, list):
        raise ValueError("JSON must be an array of objects or a wrapped array.")

    # ── Convert to DataFrame → CSV ─────────────────
    df = pd.json_normalize(data)  # flattens nested objects
    df.columns = [sanitize_name(c) for c in df.columns]

    return df.to_csv(index=False, sep=separator)


def json_to_excel(source, sheet_name: str = "Sheet1") -> bytes:
    """
    Convert JSON (file, raw text, list, or dict) → Excel bytes.

    Args:
        source     : uploaded file object | raw JSON string | list | dict
        sheet_name : name of the Excel sheet (default: Sheet1)

    Returns:
        Raw bytes of the generated .xlsx file
    """
    # ── Parse input ───────────────────────────────
    if isinstance(source, (list, dict)):
        data = source  # already parsed
    elif isinstance(source, str):
        data = json.loads(source)  # raw JSON string
    else:
        raw = source.read().decode("utf-8")  # file object
        data = json.loads(raw)

    # ── Handle all JSON shapes ─────────────────────
    if isinstance(data, dict):
        # { "data": [{...}] }  → unwrap
        for key, val in data.items():
            if isinstance(val, list):
                data = val
                break
        else:
            data = [data]  # single object → wrap in list

    if not isinstance(data, list):
        raise ValueError("JSON must be an array of objects or a wrapped array.")

    # ── Flatten nested objects ─────────────────────
    df = pd.json_normalize(data)
    df.columns = [sanitize_name(c) for c in df.columns]

    # ── Write to Excel in memory ───────────────────
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        # ── Auto-fit column widths ─────────────────
        worksheet = writer.sheets[sheet_name]
        for col_cells in worksheet.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            col_letter = col_cells[0].column_letter
            worksheet.column_dimensions[col_letter].width = min(max_len + 4, 50)

        # ── Style header row ───────────────────────
        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
        header_align = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    buffer.seek(0)
    return buffer.read()


def json_to_excel_multisheets(sources: dict) -> bytes:
    """
    Convert multiple JSON arrays → multi-sheet Excel file.

    Args:
        sources : { 'SheetName': list_or_dict, ... }

    Returns:
        Raw bytes of the generated .xlsx file
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet, data in sources.items():

            # ── Handle shapes ──────────────────────
            if isinstance(data, dict):
                for key, val in data.items():
                    if isinstance(val, list):
                        data = val
                        break
                else:
                    data = [data]

            df = pd.json_normalize(data)
            df.columns = [sanitize_name(c) for c in df.columns]
            df.to_excel(
                writer, index=False, sheet_name=sheet[:31]
            )  # Excel max 31 chars

            # ── Auto-fit + style header ────────────
            worksheet = writer.sheets[sheet[:31]]

            for col_cells in worksheet.columns:
                max_len = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col_cells
                )
                col_letter = col_cells[0].column_letter
                worksheet.column_dimensions[col_letter].width = min(max_len + 4, 50)

            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="2E75B6")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    buffer.seek(0)
    return buffer.read()


def format_json(
    source,
    indent: int = 4,
    sort_keys: bool = False,
    minify: bool = False,
    ensure_ascii: bool = False,
) -> dict:
    """
    Format / validate / minify JSON from file or raw text.

    Args:
        source       : uploaded file object OR raw JSON string
        indent       : indentation spaces (default: 4)
        sort_keys    : sort keys alphabetically (default: False)
        minify       : minify JSON (overrides indent) (default: False)
        ensure_ascii : escape non-ASCII chars (default: False)

    Returns:
        {
            'formatted'  : str,
            'is_valid'   : bool,
            'key_count'  : int,
            'size_original' : int,
            'size_formatted': int,
            'type'       : 'object' | 'array' | 'other',
            'depth'      : int,
        }
    """
    # ── Read source ───────────────────────────────
    if hasattr(source, "read"):
        raw = source.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8")
    else:
        raw = source

    raw = raw.strip()

    if not raw:
        raise ValueError("Empty input. Provide a valid JSON string or file.")

    # ── Parse JSON ────────────────────────────────
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as e:
        return {
            "formatted": None,
            "is_valid": False,
            "error": str(e),
            "error_line": e.lineno,
            "error_column": e.colno,
            "error_position": e.pos,
            "size_original": len(raw),
        }

    # ── Format ────────────────────────────────────
    if minify:
        formatted = json.dumps(
            parsed,
            separators=(",", ":"),
            ensure_ascii=ensure_ascii,
        )
    else:
        formatted = json.dumps(
            parsed,
            indent=indent,
            sort_keys=sort_keys,
            ensure_ascii=ensure_ascii,
        )

    # ── Analyze structure ─────────────────────────
    def get_depth(obj, level=0):
        if isinstance(obj, dict):
            if not obj:
                return level
            return max(get_depth(v, level + 1) for v in obj.values())
        if isinstance(obj, list):
            if not obj:
                return level
            return max(get_depth(i, level + 1) for i in obj)
        return level

    def count_keys(obj):
        if isinstance(obj, dict):
            return len(obj) + sum(count_keys(v) for v in obj.values())
        if isinstance(obj, list):
            return sum(count_keys(i) for i in obj)
        return 0

    json_type = (
        "object"
        if isinstance(parsed, dict)
        else "array" if isinstance(parsed, list) else "other"
    )

    return {
        "formatted": formatted,
        "is_valid": True,
        "type": json_type,
        "depth": get_depth(parsed),
        "key_count": count_keys(parsed),
        "item_count": len(parsed) if isinstance(parsed, (dict, list)) else 1,
        "size_original": len(raw),
        "size_formatted": len(formatted),
        "size_original_kb": round(len(raw) / 1024, 2),
        "size_formatted_kb": round(len(formatted) / 1024, 2),
        "minified": minify,
        "sorted_keys": sort_keys,
        "indent": indent,
    }


def json_to_yaml(
    source,
    indent: int = 2,
    sort_keys: bool = False,
    allow_unicode: bool = True,
    default_flow: bool = False,
    encoding: str = "utf-8",
) -> dict:
    """
    Convert JSON (file, text, dict, list) → YAML string.

    Args:
        source        : file object | raw JSON string | dict | list
        indent        : YAML indentation spaces        (default: 2)
        sort_keys     : sort keys alphabetically       (default: False)
        allow_unicode : allow unicode characters       (default: True)
        default_flow  : use flow style                 (default: False)
        encoding      : output encoding               (default: utf-8)

    Returns:
        {
            'yaml'         : str,
            'input_type'   : str,
            'key_count'    : int,
            'depth'        : int,
            'size_original': int,
            'size_yaml'    : int,
        }
    """
    import yaml

    # ── Read source ───────────────────────────────
    input_type = "text"

    if hasattr(source, "read"):
        raw = source.read()
        input_type = "file"
        if isinstance(raw, bytes):
            raw = raw.decode(encoding, errors="replace")
        data = json.loads(raw)

    elif isinstance(source, (dict, list)):
        data = source
        input_type = "object"
        raw = json.dumps(source)

    elif isinstance(source, str):
        raw = source.strip()
        data = json.loads(raw)

    elif isinstance(source, bytes):
        raw = source.decode(encoding, errors="replace")
        input_type = "bytes"
        data = json.loads(raw)

    else:
        raise ValueError("source must be a string, file, bytes, dict, or list.")

    if data is None:
        raise ValueError("Input is null/None — nothing to convert.")

    # ── Convert to YAML ───────────────────────────
    yaml_str = yaml.dump(
        data,
        indent=indent,
        sort_keys=sort_keys,
        allow_unicode=allow_unicode,
        default_flow_style=default_flow,
        explicit_start=False,
        explicit_end=False,
    )

    # ── Analyze structure ─────────────────────────
    def get_depth(obj, level=0):
        if isinstance(obj, dict):
            if not obj:
                return level
            return max(get_depth(v, level + 1) for v in obj.values())
        if isinstance(obj, list):
            if not obj:
                return level
            return max(get_depth(i, level + 1) for i in obj)
        return level

    def count_keys(obj):
        if isinstance(obj, dict):
            return len(obj) + sum(count_keys(v) for v in obj.values())
        if isinstance(obj, list):
            return sum(count_keys(i) for i in obj)
        return 0

    return {
        "yaml": yaml_str,
        "input_type": input_type,
        "type": (
            "object"
            if isinstance(data, dict)
            else "array" if isinstance(data, list) else "other"
        ),
        "key_count": count_keys(data),
        "item_count": len(data) if isinstance(data, (dict, list)) else 1,
        "depth": get_depth(data),
        "size_original": len(raw),
        "size_yaml": len(yaml_str),
        "size_original_kb": round(len(raw) / 1024, 2),
        "size_yaml_kb": round(len(yaml_str) / 1024, 2),
        "sort_keys": sort_keys,
        "indent": indent,
        "encoding": encoding,
    }


def yaml_to_json(
    source,
    indent: int = 4,
    sort_keys: bool = False,
    ensure_ascii: bool = False,
    encoding: str = "utf-8",
) -> dict:
    """
    Convert YAML (file or text) → JSON string.

    Args:
        source       : file object | raw YAML string | bytes
        indent       : JSON indentation spaces       (default: 4)
        sort_keys    : sort keys alphabetically      (default: False)
        ensure_ascii : escape non-ASCII characters   (default: False)
        encoding     : input encoding               (default: utf-8)

    Returns:
        {
            'json'          : str,
            'parsed'        : dict | list,
            'input_type'    : str,
            'type'          : str,
            'key_count'     : int,
            'depth'         : int,
            'size_original' : int,
            'size_json'     : int,
            'documents'     : int,
        }
    """
    import yaml

    # ── Read source ───────────────────────────────
    input_type = "text"

    if hasattr(source, "read"):
        raw = source.read()
        input_type = "file"
        if isinstance(raw, bytes):
            raw = raw.decode(encoding, errors="replace")
    elif isinstance(source, bytes):
        raw = source.decode(encoding, errors="replace")
        input_type = "bytes"
    elif isinstance(source, str):
        raw = source.strip()
        input_type = "text"
    else:
        raise ValueError("source must be a string, bytes, or file object.")

    if not raw.strip():
        raise ValueError("Empty input.")

    # ── Parse YAML ────────────────────────────────
    try:
        # Load all documents (handles multi-document YAML)
        documents = list(yaml.safe_load_all(raw))
    except yaml.YAMLError as e:
        raise ValueError(f"Invalid YAML: {e}")

    if not documents:
        raise ValueError("No YAML documents found.")

    # ── Single vs multi-document ──────────────────
    if len(documents) == 1:
        data = documents[0]
    else:
        # Multiple YAML docs → JSON array
        data = documents

    if data is None:
        raise ValueError("YAML parsed to null — nothing to convert.")

    # ── Convert to JSON ───────────────────────────
    json_str = json.dumps(
        data,
        indent=indent,
        sort_keys=sort_keys,
        ensure_ascii=ensure_ascii,
        default=str,  # fallback for non-serializable types
    )

    # ── Analyze ───────────────────────────────────
    def get_depth(obj, level=0):
        if isinstance(obj, dict):
            if not obj:
                return level
            return max(get_depth(v, level + 1) for v in obj.values())
        if isinstance(obj, list):
            if not obj:
                return level
            return max(get_depth(i, level + 1) for i in obj)
        return level

    def count_keys(obj):
        if isinstance(obj, dict):
            return len(obj) + sum(count_keys(v) for v in obj.values())
        if isinstance(obj, list):
            return sum(count_keys(i) for i in obj)
        return 0

    json_type = (
        "object"
        if isinstance(data, dict)
        else "array" if isinstance(data, list) else "other"
    )

    return {
        "json": json_str,
        "parsed": data,
        "input_type": input_type,
        "type": json_type,
        "key_count": count_keys(data),
        "item_count": len(data) if isinstance(data, (dict, list)) else 1,
        "depth": get_depth(data),
        "documents": len(documents),
        "size_original": len(raw),
        "size_json": len(json_str),
        "size_original_kb": round(len(raw) / 1024, 2),
        "size_json_kb": round(len(json_str) / 1024, 2),
        "sort_keys": sort_keys,
        "indent": indent,
        "encoding": encoding,
    }


def json_to_xml(
    source,
    root_element: str = "root",
    item_element: str = "item",
    indent: int = 2,
    encoding: str = "utf-8",
    include_declaration: bool = True,
    prettify: bool = True,
    attributes_mode: bool = False,
) -> dict:
    """
    Convert JSON (file, text, dict, list) → XML string.

    Args:
        source              : file object | raw JSON string | dict | list
        root_element        : root XML tag name           (default: root)
        item_element        : tag for array items         (default: item)
        indent              : XML indentation spaces      (default: 2)
        encoding            : output encoding            (default: utf-8)
        include_declaration : include <?xml ...?>         (default: True)
        prettify            : pretty print XML            (default: True)
        attributes_mode     : use attributes instead of
                              child elements for scalars  (default: False)

    Returns:
        {
            'xml'           : str,
            'type'          : str,
            'key_count'     : int,
            'depth'         : int,
            'size_original' : int,
            'size_xml'      : int,
        }
    """
    import xml.etree.ElementTree as ET
    from xml.dom import minidom
    import re

    # ── Read source ───────────────────────────────
    input_size = 0

    if hasattr(source, "read"):
        raw = source.read()
        if isinstance(raw, bytes):
            raw = raw.decode(encoding, errors="replace")
        data = json.loads(raw)
        input_size = len(raw)

    elif isinstance(source, (dict, list)):
        data = source
        input_size = len(json.dumps(source))

    elif isinstance(source, str):
        raw = source.strip()
        data = json.loads(raw)
        input_size = len(raw)

    elif isinstance(source, bytes):
        raw = source.decode(encoding, errors="replace")
        data = json.loads(raw)
        input_size = len(raw)

    else:
        raise ValueError("source must be a string, bytes, file, dict, or list.")

    # ── Validate tag names ─────────────────────────
    def sanitize_tag(name: str) -> str:
        """Make a valid XML tag name."""
        name = re.sub(r"[^\w.-]", "_", str(name))
        if name and name[0].isdigit():
            name = "_" + name
        if not name:
            name = "_"
        return name

    root_element = sanitize_tag(root_element) or "root"
    item_element = sanitize_tag(item_element) or "item"

    # ── Build XML tree ────────────────────────────
    def build_element(parent, data, tag: str):
        """Recursively build XML elements from JSON data."""
        tag = sanitize_tag(tag)

        if data is None:
            el = ET.SubElement(parent, tag)
            el.set("nil", "true")

        elif isinstance(data, bool):
            el = ET.SubElement(parent, tag)
            el.text = "true" if data else "false"
            el.set("type", "boolean")

        elif isinstance(data, int):
            el = ET.SubElement(parent, tag)
            el.text = str(data)
            el.set("type", "integer")

        elif isinstance(data, float):
            el = ET.SubElement(parent, tag)
            el.text = str(data)
            el.set("type", "float")

        elif isinstance(data, str):
            el = ET.SubElement(parent, tag)
            el.text = data

        elif isinstance(data, dict):
            el = ET.SubElement(parent, tag)
            if attributes_mode:
                # Scalar values as attributes
                children = {}
                for k, v in data.items():
                    if isinstance(v, (str, int, float, bool)) and v is not None:
                        el.set(sanitize_tag(k), str(v))
                    else:
                        children[k] = v
                for k, v in children.items():
                    build_element(el, v, k)
            else:
                for k, v in data.items():
                    build_element(el, v, k)

        elif isinstance(data, list):
            el = ET.SubElement(parent, tag)
            el.set("type", "array")
            el.set("count", str(len(data)))
            for idx, item in enumerate(data):
                # Use key name + index or item_element
                child_tag = f"{item_element}"
                build_element(el, item, child_tag)

        else:
            el = ET.SubElement(parent, tag)
            el.text = str(data)

        return el

    # ── Create root element ────────────────────────
    if isinstance(data, dict):
        root = ET.Element(root_element)
        for key, value in data.items():
            build_element(root, value, key)
    elif isinstance(data, list):
        root = ET.Element(root_element)
        root.set("type", "array")
        root.set("count", str(len(data)))
        for item in data:
            build_element(root, item, item_element)
    else:
        root = ET.Element(root_element)
        root.text = str(data)

    # ── Serialize to string ────────────────────────
    if prettify:
        xml_str = _prettify_xml(root, indent=indent)
    else:
        xml_str = ET.tostring(root, encoding="unicode")

    # ── Add XML declaration ────────────────────────
    if include_declaration:
        declaration = f'<?xml version="1.0" encoding="{encoding}"?>\n'
        xml_str = declaration + xml_str

    # ── Stats ──────────────────────────────────────
    def count_keys(obj):
        if isinstance(obj, dict):
            return len(obj) + sum(count_keys(v) for v in obj.values())
        if isinstance(obj, list):
            return sum(count_keys(i) for i in obj)
        return 0

    def get_depth(obj, level=0):
        if isinstance(obj, dict):
            if not obj:
                return level
            return max(get_depth(v, level + 1) for v in obj.values())
        if isinstance(obj, list):
            if not obj:
                return level
            return max(get_depth(i, level + 1) for i in obj)
        return level

    return {
        "xml": xml_str,
        "type": (
            "object"
            if isinstance(data, dict)
            else "array" if isinstance(data, list) else "other"
        ),
        "key_count": count_keys(data),
        "item_count": len(data) if isinstance(data, (dict, list)) else 1,
        "depth": get_depth(data),
        "size_original": input_size,
        "size_xml": len(xml_str),
        "size_original_kb": round(input_size / 1024, 2),
        "size_xml_kb": round(len(xml_str) / 1024, 2),
        "root_element": root_element,
        "item_element": item_element,
        "encoding": encoding,
    }


def _prettify_xml(element, indent: int = 2) -> str:
    """
    Pretty-print XML element with proper indentation.
    Pure Python — no minidom dependency.
    """
    import xml.etree.ElementTree as ET

    def _indent(el, level=0):
        pad = "\n" + " " * indent * level
        if len(el):
            if not el.text or not el.text.strip():
                el.text = pad + " " * indent
            if not el.tail or not el.tail.strip():
                el.tail = pad
            for child in el:
                _indent(child, level + 1)
            if not child.tail or not child.tail.strip():
                child.tail = pad
        else:
            if level and (not el.tail or not el.tail.strip()):
                el.tail = pad
        return el

    _indent(element)
    return ET.tostring(element, encoding="unicode")

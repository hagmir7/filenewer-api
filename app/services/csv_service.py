"""
CSV service functions.
"""

import io
import re
import logging

import pandas as pd

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# CSV TO SQL
# ─────────────────────────────────────────────

def sanitize_name(name: str) -> str:
    name = name.strip().lower()
    name = re.sub(r"[^a-z0-9_]", "_", name)
    if name[0].isdigit():
        name = "col_" + name
    return name


def infer_sql_type(dtype) -> str:
    dtype = str(dtype)
    if "int" in dtype:
        return "INTEGER"
    if "float" in dtype:
        return "REAL"
    if "bool" in dtype:
        return "BOOLEAN"
    if "datetime" in dtype:
        return "TIMESTAMP"
    return "TEXT"


def csv_to_sql(source, table_name: str, dialect: str = 'sqlite', separator: str = None) -> str:
    """
    Convert CSV (file or raw text string) → SQL statements.

    Args:
        source     : uploaded file object OR raw CSV string
        table_name : name for the SQL table
        dialect    : 'sqlite' | 'postgresql' | 'mysql'

    Returns:
        SQL string (CREATE TABLE + INSERT statements)
    """
    # ── Read CSV ─────────────────────────────────
    if isinstance(source, str):
        df = pd.read_csv(io.StringIO(source), sep=separator, engine='python')
    else:
        # ── File: read bytes first, then decode ──
        raw = source.read().decode('utf-8')          # ← read raw bytes
        df  = pd.read_csv(io.StringIO(raw), sep=separator, engine='python')

    df.columns = [sanitize_name(c) for c in df.columns]
    table_name = sanitize_name(table_name)

    # ── Dialect config ────────────────────────────
    if dialect == "postgresql":
        pk = "id SERIAL PRIMARY KEY"
        str_quote = "E'{}'"
    elif dialect == "mysql":
        pk = "id INT AUTO_INCREMENT PRIMARY KEY"
        str_quote = "'{}'"
    else:  # sqlite (default)
        pk = "id INTEGER PRIMARY KEY AUTOINCREMENT"
        str_quote = "'{}'"

    # ── CREATE TABLE ──────────────────────────────
    col_defs = ",\n  ".join(
        f'"{col}" {infer_sql_type(dtype)}' for col, dtype in df.dtypes.items()
    )
    sql_lines = [
        f"-- Generated SQL ({dialect}) for table: {table_name}",
        f"-- Rows: {len(df)}  |  Columns: {len(df.columns)}\n",
        f'CREATE TABLE IF NOT EXISTS "{table_name}" (',
        f"  {pk},",
        f"  {col_defs}",
        ");\n",
    ]

    # ── INSERT statements ─────────────────────────
    col_str = ", ".join(f'"{c}"' for c in df.columns)

    for _, row in df.iterrows():
        values = []
        for val in row:
            if pd.isna(val):
                values.append("NULL")
            elif isinstance(val, bool):
                values.append("TRUE" if val else "FALSE")
            elif isinstance(val, (int, float)):
                values.append(str(val))
            else:
                escaped = str(val).replace("'", "''")
                values.append(f"'{escaped}'")

        val_str = ", ".join(values)
        sql_lines.append(f'INSERT INTO "{table_name}" ({col_str}) VALUES ({val_str});')

    return "\n".join(sql_lines)


# ─────────────────────────────────────────────
# CSV TO JSON
# ─────────────────────────────────────────────
def csv_to_json(source, separator: str = None, orient: str = "records") -> list | dict:
    """
    Convert CSV (file or raw text string) → JSON.

    Args:
        source    : uploaded file object OR raw CSV string
        separator : column separator (None = auto-detect)
        orient    : 'records'  → [ {col: val}, ... ]         (default)
                    'columns'  → { col: {index: val}, ... }
                    'values'   → [ [val, val], ... ]
                    'index'    → { index: {col: val}, ... }

    Returns:
        parsed Python object (list or dict) ready for JSON response
    """
    # ── Read CSV ──────────────────────────────────
    if isinstance(source, str):
        df = pd.read_csv(io.StringIO(source), sep=separator, engine="python")
    else:
        raw = source.read().decode("utf-8")
        df = pd.read_csv(io.StringIO(raw), sep=separator, engine="python")

    # ── Clean column names ─────────────────────────
    df.columns = [sanitize_name(c) for c in df.columns]

    # ── Replace NaN with None (JSON null) ─────────
    df = df.where(pd.notnull(df), None)

    # ── Convert to JSON-serializable object ────────
    return df.to_dict(orient=orient)


def csv_to_excel(source, sheet_name: str = "Sheet1", separator: str = None) -> bytes:
    """
    Convert CSV (file or raw text string) → Excel bytes.

    Args:
        source     : uploaded file object OR raw CSV string
        sheet_name : name of the Excel sheet (default: Sheet1)
        separator  : column separator (None = auto-detect)

    Returns:
        Raw bytes of the generated .xlsx file
    """
    # ── Read CSV ──────────────────────────────────
    if isinstance(source, str):
        df = pd.read_csv(io.StringIO(source), sep=separator, engine="python")
    else:
        raw = source.read().decode("utf-8")
        df = pd.read_csv(io.StringIO(raw), sep=separator, engine="python")

    # ── Clean column names ─────────────────────────
    df.columns = [sanitize_name(str(c)) for c in df.columns]

    # ── Replace NaN with empty string ─────────────
    df = df.fillna("")

    # ── Write to Excel in memory ───────────────────
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        worksheet = writer.sheets[sheet_name]

        # ── Auto-fit column widths ─────────────────
        for col_cells in worksheet.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in col_cells
            )
            col_letter = col_cells[0].column_letter
            worksheet.column_dimensions[col_letter].width = min(max_len + 4, 50)

        # ── Style header row ───────────────────────
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # ── Style data rows (alternating colors) ───
        from openpyxl.styles import PatternFill as PF

        even_fill = PF(fill_type="solid", fgColor="DCE6F1")  # light blue
        odd_fill = PF(fill_type="solid", fgColor="FFFFFF")  # white

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            fill = even_fill if row_idx % 2 == 0 else odd_fill
            for cell in row:
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # ── Freeze header row ──────────────────────
        worksheet.freeze_panes = "A2"

        # ── Set row height for header ──────────────
        worksheet.row_dimensions[1].height = 30

    buffer.seek(0)
    return buffer.read()


def csv_to_excel_multisheets(sources: dict, separator: str = None) -> bytes:
    """
    Convert multiple CSV strings → multi-sheet Excel file.

    Args:
        sources   : { 'SheetName': 'csv string...', ... }
        separator : column separator (None = auto-detect)

    Returns:
        Raw bytes of the generated .xlsx file
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet, csv_data in sources.items():

            # ── Read CSV ───────────────────────────
            if isinstance(csv_data, str):
                df = pd.read_csv(io.StringIO(csv_data), sep=separator, engine="python")
            else:
                raw = csv_data.read().decode("utf-8")
                df = pd.read_csv(io.StringIO(raw), sep=separator, engine="python")

            df.columns = [sanitize_name(str(c)) for c in df.columns]
            df = df.fillna("")

            sheet_label = sheet[:31]  # Excel max 31 chars
            df.to_excel(writer, index=False, sheet_name=sheet_label)

            worksheet = writer.sheets[sheet_label]

            # ── Auto-fit columns ───────────────────
            for col_cells in worksheet.columns:
                max_len = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in col_cells
                )
                col_letter = col_cells[0].column_letter
                worksheet.column_dimensions[col_letter].width = min(max_len + 4, 50)

            # ── Style header ───────────────────────
            thin = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(fill_type="solid", fgColor="2E75B6")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin

            # ── Alternating row colors ─────────────
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                fill_color = "DCE6F1" if row_idx % 2 == 0 else "FFFFFF"
                for cell in row:
                    cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
                    cell.border = thin
                    cell.alignment = Alignment(vertical="center")

            worksheet.freeze_panes = "A2"
            worksheet.row_dimensions[1].height = 30

    buffer.seek(0)
    return buffer.read()


def view_csv(
    source,
    separator: str = None,
    encoding: str = "utf-8",
    max_rows: int = 1000,
    page: int = 1,
    page_size: int = 100,
    sort_by: str = None,
    sort_order: str = "asc",
    filter_column: str = None,
    filter_value: str = None,
    filter_operator: str = "contains",
    columns: list = None,
) -> dict:
    """
    Parse and view CSV file with pagination, sorting, and filtering.

    Args:
        source          : uploaded file object OR raw text string
        separator       : column separator (None = auto-detect)
        encoding        : file encoding                  (default: utf-8)
        max_rows        : max rows to process            (default: 1000)
        page            : page number (1-based)          (default: 1)
        page_size       : rows per page                  (default: 100)
        sort_by         : column name to sort by         (default: None)
        sort_order      : 'asc' | 'desc'                 (default: asc)
        filter_column   : column to filter on            (default: None)
        filter_value    : value to filter by             (default: None)
        filter_operator : 'contains' | 'equals'
                          'starts_with' | 'ends_with'
                          'greater_than' | 'less_than'
                          'not_empty' | 'is_empty'       (default: contains)
        columns         : list of columns to return      (default: all)

    Returns:
        {
            'columns'      : list,
            'rows'         : list,
            'total_rows'   : int,
            'total_columns': int,
            'page'         : int,
            'page_size'    : int,
            'total_pages'  : int,
            'stats'        : dict,
        }
    """
    import math

    # ── Read source ───────────────────────────────
    if hasattr(source, "read"):
        raw = source.read()
        if isinstance(raw, bytes):
            raw = raw.decode(encoding, errors="replace")
    elif isinstance(source, str):
        raw = source
    else:
        raise ValueError("source must be a string or file object.")

    if not raw.strip():
        raise ValueError("Empty input.")

    # ── Parse CSV ────────────────────────────────
    df = pd.read_csv(
        io.StringIO(raw),
        sep=separator,
        engine="python",
        nrows=max_rows,
        dtype=str,  # keep everything as string for display
        keep_default_na=False,  # don't convert empty to NaN
    )

    # ── Clean column names ─────────────────────────
    df.columns = [str(c).strip() for c in df.columns]

    # ── Select columns ─────────────────────────────
    if columns:
        valid_cols = [c for c in columns if c in df.columns]
        if valid_cols:
            df = df[valid_cols]

    total_rows = len(df)
    total_columns = len(df.columns)
    all_columns = list(df.columns)

    # ── Apply filter ──────────────────────────────
    if filter_column and filter_column in df.columns and filter_value is not None:
        col = df[filter_column].astype(str)

        if filter_operator == "contains":
            df = df[col.str.contains(filter_value, case=False, na=False)]
        elif filter_operator == "equals":
            df = df[col.str.lower() == filter_value.lower()]
        elif filter_operator == "starts_with":
            df = df[col.str.lower().str.startswith(filter_value.lower())]
        elif filter_operator == "ends_with":
            df = df[col.str.lower().str.endswith(filter_value.lower())]
        elif filter_operator == "greater_than":
            try:
                df = df[pd.to_numeric(col, errors="coerce") > float(filter_value)]
            except Exception:
                pass
        elif filter_operator == "less_than":
            try:
                df = df[pd.to_numeric(col, errors="coerce") < float(filter_value)]
            except Exception:
                pass
        elif filter_operator == "not_empty":
            df = df[col.str.strip() != ""]
        elif filter_operator == "is_empty":
            df = df[col.str.strip() == ""]

    filtered_rows = len(df)

    # ── Apply sort ────────────────────────────────
    if sort_by and sort_by in df.columns:
        ascending = sort_order.lower() != "desc"
        try:
            # Try numeric sort first
            df = df.sort_values(
                by=sort_by,
                ascending=ascending,
                key=lambda x: pd.to_numeric(x, errors="coerce"),
                na_position="last",
            )
        except Exception:
            df = df.sort_values(
                by=sort_by,
                ascending=ascending,
                na_position="last",
            )

    # ── Pagination ────────────────────────────────
    total_pages = max(1, math.ceil(filtered_rows / page_size))
    page = max(1, min(page, total_pages))
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    page_df = df.iloc[start_idx:end_idx]

    # ── Build rows ────────────────────────────────
    rows = page_df.to_dict(orient="records")

    # ── Column stats ──────────────────────────────
    col_stats = {}
    for col in df.columns:
        series = df[col]
        numeric = pd.to_numeric(series, errors="coerce")
        is_numeric = numeric.notna().sum() > len(series) * 0.5

        stat = {
            "name": col,
            "type": "numeric" if is_numeric else "text",
            "empty_count": (series.str.strip() == "").sum(),
            "unique": series.nunique(),
        }

        if is_numeric:
            stat.update(
                {
                    "min": (
                        round(float(numeric.min()), 4)
                        if not numeric.isna().all()
                        else None
                    ),
                    "max": (
                        round(float(numeric.max()), 4)
                        if not numeric.isna().all()
                        else None
                    ),
                    "mean": (
                        round(float(numeric.mean()), 4)
                        if not numeric.isna().all()
                        else None
                    ),
                    "sum": (
                        round(float(numeric.sum()), 4)
                        if not numeric.isna().all()
                        else None
                    ),
                }
            )

        col_stats[col] = stat

    return {
        "columns": all_columns,
        "rows": rows,
        "total_rows": total_rows,
        "filtered_rows": filtered_rows,
        "total_columns": total_columns,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "has_next": page < total_pages,
        "has_prev": page > 1,
        "separator": separator or "auto",
        "encoding": encoding,
        "col_stats": col_stats,
        "filter": (
            {
                "column": filter_column,
                "value": filter_value,
                "operator": filter_operator,
            }
            if filter_column
            else None
        ),
        "sort": (
            {
                "column": sort_by,
                "order": sort_order,
            }
            if sort_by
            else None
        ),
    }
